from flask import Blueprint, jsonify, request, send_file
import json
from openpyxl import Workbook
from flask import Blueprint, jsonify, request, send_file, current_app
import requests
from requests.auth import HTTPBasicAuth

invoice_template_bp = Blueprint("invoice_template_bp", __name__)

FIELDS_FILE = "invoice_template_fields.json"


# ===============================
# GET ALL INVOICE FIELDS
# ===============================
@invoice_template_bp.route("/invoice-template/fields", methods=["GET"])
def get_invoice_template_fields():
    with open(FIELDS_FILE, "r") as f:
        return jsonify(json.load(f))
# ===============================
# GET OTM INVOICE METADATA
# ===============================
@invoice_template_bp.route("/invoice-template/otm-metadata", methods=["GET"])
def get_otm_invoice_metadata():
    url = current_app.config["OTM_INVOICE_METADATA_URL"]

    response = requests.get(
        url,
        auth=HTTPBasicAuth(
            current_app.config["OTM_USERNAME"],
            current_app.config["OTM_PASSWORD"]
        ),
        headers={"Accept": "application/json"}
    )

    return jsonify(response.json()), response.status_code



# ===============================
# GENERATE EXCEL TEMPLATE
# ===============================
@invoice_template_bp.route("/invoice-template/generate", methods=["POST"])
def generate_invoice_template():
    data = request.json

    wb = Workbook()

    # -------- HEADER SHEET --------
    header_sheet = wb.active
    header_sheet.title = "Invoice_Header"
    header_sheet.append(data.get("header", []))

    # -------- REFNUM SHEET --------
    ref_sheet = wb.create_sheet("Invoice_Reference_Numbers")
    ref_sheet.append(["Invoice ID"] + data.get("referenceNumbers", []))

    # -------- LINE ITEMS SHEET --------
    line_sheet = wb.create_sheet("Invoice_Line_Items")
    line_sheet.append(["Invoice ID"] + data.get("lineItems", []))

    file_path = "invoice_template.xlsx"
    wb.save(file_path)

    return send_file(file_path, as_attachment=True)


# ===============================
# TEMPLATE LIST & SAVE (JSON)
# ===============================
TEMPLATES_FILE = "invoice_templates.json"
import os
from uuid import uuid4

@invoice_template_bp.route("/invoice-template/templates", methods=["GET"])
def get_templates():
    if not os.path.exists(TEMPLATES_FILE):
        return jsonify([])
    
    try:
        with open(TEMPLATES_FILE, "r") as f:
            templates = json.load(f)
            return jsonify(templates)
    except Exception:
        return jsonify([])

@invoice_template_bp.route("/invoice-template/templates", methods=["POST"])
def save_template():
    # Expecting: { "name": "My Template", "fields": [...] }
    data = request.json
    name = data.get("name")
    fields = data.get("fields")
    
    if not name or not fields:
        return jsonify({"error": "Missing name or fields"}), 400

    # Load existing
    templates = []
    if os.path.exists(TEMPLATES_FILE):
        try:
            with open(TEMPLATES_FILE, "r") as f:
                templates = json.load(f)
        except:
            templates = []

    # New Template
    new_template = {
        "id": str(uuid4()),
        "name": name,
        "fields": fields,
        # "createdAt": ... 
    }
    
    templates.append(new_template)
    
    with open(TEMPLATES_FILE, "w") as f:
        json.dump(templates, f, indent=2)

    return jsonify({"message": "Template saved", "id": new_template["id"]})
